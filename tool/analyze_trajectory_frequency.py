#!/usr/bin/env python3
"""
Analyze trajectory message timestamps (JSONL) to compute generation frequency.

Outputs:
  - Excel file with time/delta/frequency, in-sheet statistics, and an embedded chart.

Optional extras (to illustrate further analysis ideas):
  --histogram saves a frequency histogram PNG.
  --moving-average-window computes a moving-average column.
  --detect-outliers flags points beyond mean ± 3σ in the Excel output.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NS_TO_SEC = 1e-9


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Trajectory frequency analyzer")
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Path to trajectory JSONL file (each line contains a timestamp field)",
    )
    parser.add_argument(
        "--scenario",
        required=True,
        type=int,
        choices=[1, 2, 3],
        help="Scenario number (1-3)",
    )
    parser.add_argument(
        "--histogram",
        action="store_true",
        help="Generate a histogram of frequency values (<prefix>_histogram.png)",
    )
    parser.add_argument(
        "--moving-average-window",
        type=int,
        default=0,
        help="Window size for moving average overlay (0 to disable)",
    )
    parser.add_argument(
        "--detect-outliers",
        action="store_true",
        help="Mark frequency outliers beyond mean ± 3σ in the Excel output",
    )
    return parser.parse_args()


def extract_timestamp_ns(record: Dict) -> int:
    """
    Extract timestamp in nanoseconds from a JSONL record.

    Primary expectation: a top-level 'timestamp' field (int, nanoseconds).
    Fallback: header.stamp.sec/nanosec within a 'message' field.
    """
    if "timestamp" in record:
        try:
            return int(record["timestamp"])
        except (TypeError, ValueError) as exc:
            raise ValueError("Invalid timestamp value in record") from exc

    msg = record.get("message", {})
    header = msg.get("header", {})
    stamp = header.get("stamp")
    if isinstance(stamp, dict) and "sec" in stamp and "nanosec" in stamp:
        sec = int(stamp.get("sec", 0))
        nsec = int(stamp.get("nanosec", 0))
        return sec * 1_000_000_000 + nsec

    raise ValueError("No timestamp found in record")


def load_timestamps(path: Path) -> List[int]:
    """Load all timestamps (nanoseconds) from a JSONL file."""
    timestamps: List[int] = []
    with path.open() as f:
        for line_num, line in enumerate(f, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            try:
                record = json.loads(stripped)
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSON on line {line_num}") from exc
            ts = extract_timestamp_ns(record)
            timestamps.append(ts)

    if not timestamps:
        raise ValueError("No timestamps found in the provided file.")
    return timestamps


def build_frequency_dataframe(timestamps_ns: Iterable[int]) -> pd.DataFrame:
    """Compute time since start, delta times, and frequencies."""
    ts_array = np.asarray(list(timestamps_ns), dtype=np.int64)
    timestamp_sec = ts_array.astype(np.float64) * NS_TO_SEC
    kst_dt = pd.to_datetime(ts_array, unit="ns", utc=True).tz_convert("Asia/Seoul")
    kst_nsec = pd.Series(ts_array % 1_000_000_000, dtype="int64")
    kst_time = kst_dt.strftime("%Y-%m-%d %H:%M:%S.") + kst_nsec.astype(str).str.zfill(9)
    t0 = ts_array[0]
    time_sec = (ts_array - t0) * NS_TO_SEC

    delta_sec = np.full_like(time_sec, fill_value=np.nan, dtype=np.float64)
    delta_sec[1:] = (ts_array[1:] - ts_array[:-1]) * NS_TO_SEC

    freq_hz = np.full_like(time_sec, fill_value=np.nan, dtype=np.float64)
    with np.errstate(divide="ignore", invalid="ignore"):
        positive_delta = delta_sec[1:] > 0
        freq_hz[1:][positive_delta] = 1.0 / delta_sec[1:][positive_delta]

    df = pd.DataFrame(
        {
            "Unix Time(Timestamp) [sec]": timestamp_sec,
            "KST Time": kst_time,
            "Time [sec]": time_sec,
            "Delta Time [sec]": delta_sec,
            "Frequency [Hz]": freq_hz,
        }
    )
    return df


def compute_statistics(freq_series: pd.Series) -> Dict[str, float]:
    """Calculate mean, std, and variance over valid frequencies."""
    valid = freq_series.dropna()
    if valid.empty:
        return {
            k: float("nan") for k in ["Mean [Hz]", "Std [Hz]", "Variance [Hz^2]"]
        }

    mean = float(valid.mean())
    variance = float(valid.var(ddof=0))
    std = float(np.sqrt(variance))
    return {
        "Mean [Hz]": mean,
        "Std [Hz]": std,
        "Variance [Hz^2]": variance,
    }


def detect_outliers(freq_series: pd.Series, mean: float, std: float, sigma: float = 3.0) -> pd.Series:
    """
    Flag outliers beyond mean ± sigma*std.
    Returns a boolean Series aligned with freq_series.
    """
    if np.isnan(std) or std == 0.0 or freq_series.dropna().empty:
        return pd.Series([False] * len(freq_series), index=freq_series.index)
    lower = mean - sigma * std
    upper = mean + sigma * std
    return (freq_series < lower) | (freq_series > upper)


def add_moving_average(freq_series: pd.Series, window: int) -> pd.Series:
    """Compute a centered moving average; retains NaNs where no data exists."""
    return freq_series.rolling(window=window, min_periods=1, center=True).mean()


def plot_histogram(freq_series: pd.Series, output_png: Path, bins: int = 40) -> None:
    """Plot and save a histogram of frequency values (excluding NaNs)."""
    valid = freq_series.dropna()
    if valid.empty:
        return
    plt.figure(figsize=(8, 4))
    plt.hist(valid, bins=bins, color="tab:blue", edgecolor="black", alpha=0.75)
    plt.xlabel("Trajectory Frequency [Hz]")
    plt.ylabel("Count")
    plt.title("Trajectory Frequency Histogram")
    plt.grid(True, linestyle="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(output_png, dpi=200)
    plt.close()


def cumulative_mean(series: pd.Series) -> pd.Series:
    """Compute cumulative mean, ignoring leading NaNs and propagating NaN when current value is NaN."""
    values = series.values
    cum_sum = 0.0
    count = 0
    result: List[float] = []
    for val in values:
        if np.isnan(val):
            result.append(np.nan)
            continue
        cum_sum += val
        count += 1
        result.append(cum_sum / count if count > 0 else np.nan)
    return pd.Series(result, index=series.index, name="Cumulative Mean [Hz]")


def write_excel(
    output_path: Path,
    raw_df: pd.DataFrame,
    stats: Dict[str, float],
    scenario: int,
    include_outliers: bool,
    input_name: str,
) -> None:
    """
    Write data and statistics to a single Excel sheet with styles and embedded chart.
    Statistics layout:
      I1:J1: header (Value in J1)
      I2/J2: Trajectory 생성 주기 (평균) [Hz]
      I3/J3: 표준편차 [Hz]
      I4/J4: 분산 [Hz²]
    Chart inserted at I7 showing mean over time.
    """
    sheet_name = "TrajectoryFrequency"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Prepare export headers per requirements
        export_df = raw_df.copy()
        export_df = export_df.rename(
            columns={
                "Time [sec]": "Time [sec]",
                "Delta Time [sec]": "Delta [sec]",
                "Frequency [Hz]": "Freq. [Hz]",
                "Cumulative Mean [Hz]": "Mean [Hz]",
            }
        )
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        sheet = writer.sheets[sheet_name]
        # 엑셀 기본 확대 배율 설정 (110%)
        sheet.sheet_view.zoomScale = 110

        # Number format: timestamp/time/delta four decimals, freq/ref/mean two decimals
        format_four = "0.0000"
        format_two = "0.00"
        numeric_four_cols = ["Unix Time(Timestamp) [sec]", "Time [sec]", "Delta Time [sec]"]
        numeric_two_cols = ["Frequency [Hz]", "Reference Hz", "Cumulative Mean [Hz]"]
        for col_name in numeric_four_cols:
            col_idx = raw_df.columns.get_loc(col_name) + 1
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = format_four
        for col_name in numeric_two_cols:
            if col_name not in raw_df.columns:
                continue
            col_idx = raw_df.columns.get_loc(col_name) + 1
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = format_two
        # Center alignment for A-G columns (including header)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # Smaller font for timestamp column (A)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(size=9)

        # Statistics labels and values in I/J (H kept empty)
        sheet["I1"] = "Statistics"
        sheet["J1"] = "Value"
        sheet["I2"] = "Trajectory 생성 주기 (평균) [Hz]"
        sheet["I3"] = "표준편차 [Hz]"
        sheet["I4"] = "분산 [Hz²]"
        sheet["J2"] = stats.get("Mean [Hz]", float("nan"))
        sheet["J3"] = stats.get("Std [Hz]", float("nan"))
        sheet["J4"] = stats.get("Variance [Hz^2]", float("nan"))

        # Styling for statistics box
        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        sheet["I1"].fill = header_fill
        sheet["J1"].fill = header_fill
        sheet["I1"].font = Font(bold=True)
        sheet["J1"].font = Font(bold=True)
        sheet["I2"].font = Font(bold=True)
        stats_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
        for row in sheet.iter_rows(min_row=2, max_row=4, min_col=9, max_col=10):
            for cell in row:
                cell.fill = stats_fill
        sheet["J2"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        sheet["J2"].font = Font(bold=True, size=12)
        sheet["J3"].font = Font(bold=False, size=12)
        sheet["J4"].font = Font(bold=False, size=12)
        for cell in sheet["J2:J4"]:
            for c in cell:
                c.number_format = format_two

        align_center = Alignment(horizontal="center", vertical="center")
        sheet["I1"].alignment = align_center
        sheet["J1"].alignment = align_center
        for row in sheet.iter_rows(min_row=2, max_row=4, min_col=9, max_col=10):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        thin = Side(border_style="thin", color="666666")
        box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in sheet.iter_rows(min_row=1, max_row=4, min_col=9, max_col=10):
            for cell in row:
                cell.border = box_border

        # Fixed column widths
        fixed_widths = {
            1: 16.0,
            2: 26.0,
            3: 10.0,
            4: 10.0,
            5: 10.0,
            6: 10.0,
            7: 10.0,
            8: 10.0 / 3.0,
            9: 28.0,
            10: 16.0,
        }
        for col_idx, width in fixed_widths.items():
            col_letter = sheet.cell(row=1, column=col_idx).column_letter
            sheet.column_dimensions[col_letter].width = width

        # Chart ranges
        row_count = len(raw_df) + 1  # +1 for header row
        time_col = raw_df.columns.get_loc("Time [sec]") + 1
        freq_col = raw_df.columns.get_loc("Frequency [Hz]") + 1
        cum_mean_col = raw_df.columns.get_loc("Cumulative Mean [Hz]") + 1
        chart = ScatterChart()
        chart.title = f"Trajectory 생성 주기 ({input_name})"
        chart.y_axis.title = "Freq. [Hz]"
        chart.x_axis.title = "Time [sec]"
        grid_line_style = LineProperties(solidFill="DDDDDD")
        grid_line_style.w = 8000
        chart.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=grid_line_style))
        chart.x_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=grid_line_style))
        chart.legend.position = "b"
        gp = chart.plot_area.graphicalProperties
        if gp is None:
            gp = GraphicalProperties()
            chart.plot_area.graphicalProperties = gp
        gp.solidFill = "FAFAFA"
        chart.width = 6.0 * 2.5 * 0.95
        chart.height = 6.0 * 2.5 * 0.75 * 0.95

        # Y-axis starts at 0; cap upper bound to keep reference line visible.
        mean_val = stats.get("Mean [Hz]")
        chart.y_axis.scaling.min = 0.0
        if mean_val is not None and not np.isnan(mean_val):
            chart.y_axis.scaling.max = max(mean_val * 1.14, 5.5)
        else:
            chart.y_axis.scaling.max = 5.5

        # Show x-axis ticks every 5 seconds (integer labels only)
        chart.x_axis.majorUnit = 5.0
        chart.x_axis.number_format = "0"

        time_ref = Reference(sheet, min_col=time_col, min_row=2, max_row=row_count)

        def add_series(col_idx: int, name: str, color: str | None, dash_style: str | None) -> None:
            data_ref = Reference(sheet, min_col=col_idx, min_row=2, max_row=row_count)
            series = Series(data_ref, xvalues=time_ref, title=name)
            series.marker.symbol = "none"
            series.graphicalProperties.line.width = 8500
            if color:
                series.graphicalProperties.line.solidFill = color
            if dash_style:
                series.graphicalProperties.line.dashStyle = dash_style
            chart.series.append(series)

        add_series(cum_mean_col, "Mean", color="00AA00", dash_style=None)
        if "Reference Hz" in raw_df.columns:
            ref_col = raw_df.columns.get_loc("Reference Hz") + 1
            add_series(ref_col, "Reference Hz", color="FF0000", dash_style="sysDash")

        sheet.add_chart(chart, "I7")

        if include_outliers and "Outlier (>|3σ|)" in raw_df.columns:
            pass  # Outliers remain in the same sheet for transparency


def main() -> None:
    args = parse_args()

    timestamps_ns = load_timestamps(args.input)
    raw_df = build_frequency_dataframe(timestamps_ns)

    # Cumulative metrics for plotting and inspection
    raw_df["Cumulative Mean [Hz]"] = cumulative_mean(raw_df["Frequency [Hz]"])
    raw_df["Reference Hz"] = 5.0

    stats = compute_statistics(raw_df["Frequency [Hz]"])
    outlier_series = None
    if args.detect_outliers:
        outlier_series = detect_outliers(
            raw_df["Frequency [Hz]"], stats["Mean [Hz]"], stats["Std [Hz]"], sigma=3.0
        )
        raw_df["Outlier (>|3σ|)"] = outlier_series

    moving_avg_series = None
    if args.moving_average_window and args.moving_average_window > 0:
        moving_avg_series = add_moving_average(raw_df["Frequency [Hz]"], args.moving_average_window)
        raw_df[f"Moving Avg (win={args.moving_average_window}) [Hz]"] = moving_avg_series

    # Ensure primary columns are ordered up front
    primary_cols = [
        "Unix Time(Timestamp) [sec]",
        "KST Time",
        "Time [sec]",
        "Delta Time [sec]",
        "Frequency [Hz]",
        "Reference Hz",
        "Cumulative Mean [Hz]",
    ]
    other_cols = [col for col in raw_df.columns if col not in primary_cols]
    raw_df = raw_df[primary_cols + other_cols]

    # Output file placed under output/ with name based on input stem for clarity.
    excel_path = Path("output") / f"{args.input.stem}_freq.xlsx"
    write_excel(
        excel_path,
        raw_df,
        stats,
        args.scenario,
        include_outliers=args.detect_outliers,
        input_name=args.input.stem,
    )

    if args.histogram:
        hist_path = Path("output") / f"{args.input.stem}_histogram.png"
        plot_histogram(raw_df["Frequency [Hz]"], hist_path)


if __name__ == "__main__":
    main()

# ---------------------------------------------------------------------------
# Additional analysis suggestions (hooked into CLI flags above):
#   - Histogram: run with --histogram to inspect distribution of frequencies.
#   - Moving average: --moving-average-window N adds a stability column.
#   - Outlier detection: --detect-outliers marks points outside mean ±3σ.
# These hooks make it easy to extend further for scenario-to-scenario comparisons.
# ---------------------------------------------------------------------------
